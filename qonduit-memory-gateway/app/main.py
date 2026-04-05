from __future__ import annotations

from fastapi import FastAPI, HTTPException, UploadFile, File, Form, Request
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
import httpx
import json
import ast
import csv
import os
import logging
import time
import asyncio
from pathlib import Path
import re
import uuid
from typing import Any

from pypdf import PdfReader
from docx import Document
from openpyxl import load_workbook

from .budget import build_budget, estimate_tokens, trim_recent_messages
from .store import load_conversation, save_conversation
from .summarizer import summarize_messages
from .rag import (
    ensure_collection,
    add_document,
    search_documents,
    list_collections,
    create_collection_marker,
    qdrant,
    COLLECTION_NAME,
)
from qdrant_client.models import Filter, FieldCondition, MatchValue

app = FastAPI(title="Qonduit Memory Gateway")
logger = logging.getLogger("qonduit.memory_gateway")


@app.on_event("startup")
async def startup() -> None:
    ensure_collection()


LLAMA_BASE = "http://192.168.5.5:8080"
UPSTREAM_CONNECT_TIMEOUT_SECONDS = 10.0
UPSTREAM_WRITE_TIMEOUT_SECONDS = 60.0
UPSTREAM_POOL_TIMEOUT_SECONDS = 60.0
STREAM_KEEPALIVE_INTERVAL_SECONDS = 2.0

UPLOAD_DIR = "/mnt/models/qonduit_uploads"

TEXT_EXTENSIONS = {
    ".txt", ".md", ".json", ".csv",
    ".py", ".c", ".cpp", ".h", ".hpp",
    ".java", ".kt", ".kts",
    ".xml", ".html", ".css", ".js", ".ts",
    ".sql", ".yaml", ".yml", ".toml", ".ini", ".sh",
    ".go", ".rs", ".swift", ".php", ".rb", ".pl", ".lua",
    ".vhdl", ".vhd", ".v", ".dart",
}

DEFAULT_SYSTEM_PROMPT = (
    "You are Qonduit, a practical local coding and systems assistant. "
    "Be accurate, structured, and concise. "
    "For any non-trivial task, think in phases and present the answer in small, bounded chunks. "

    "Core response rules: "
    "1. Break complex work into numbered steps or short sections. "
    "2. Prefer the smallest complete useful answer over a long answer. "
    "3. Do not try to fit an oversized answer into one response. "
    "4. If the full answer would be long, give the first useful chunk only, stop at a natural boundary, and end with: "
    "\"Reply with continue for the next chunk.\" "
    "5. Never ramble. Avoid repetition and unnecessary explanation. "
    "6. For coding tasks, give the plan first if the task is large, then provide only the current step's code. "
    "7. For debugging, start with the most likely cause, then the next concrete action. "
    "8. For infrastructure or setup tasks, prefer exact commands and explicit file edits. "
    "9. Keep each response self-contained and easy to apply immediately. "
    "10. If the user asks for full detail, still split the answer into manageable parts instead of one huge block. "

    "Formatting rules: "
    "Use short headings when helpful. "
    "Use short numbered steps for procedures. "
    "Keep paragraphs short. "
    "When giving code, include only code needed for the current step unless the user explicitly asks for the full file. "

    "Behavior rules: "
    "If context is large or the task is broad, summarize the plan briefly before details. "
    "If multiple valid approaches exist, recommend one and keep alternatives brief. "
    "If the answer risks being cut off, compress and stop cleanly rather than continuing mid-thought. "
    "If relevant information exists in the provided context, conversation summary, or knowledge base excerpts, use it directly instead of guessing. "
    "Optimize for reliability, clarity, and completion over verbosity."
)


class ChatMessage(BaseModel):
    role: str
    content: Any


class GatewayChatRequest(BaseModel):
    conversation_id: str | None = None
    messages: list[ChatMessage]
    model: str
    context_size: int | None = Field(default=None)
    max_tokens: int = Field(default=2048)
    temperature: float = Field(default=0.7)
    stream: bool = False
    user: str | None = None
    rag_collection: str | None = None

    model_config = {"extra": "allow"}

    def resolved_context_size(self) -> int:
        value = self.context_size or 65536
        return max(value, 1024)


class RagIngestRequest(BaseModel):
    text: str
    source: str = "manual_test"
    collection: str = "default"
    document_name: str = "untitled"


class RagSearchRequest(BaseModel):
    query: str
    limit: int = 4
    collection: str = "default"


class RagCollectionCreateRequest(BaseModel):
    name: str


class RagCollectionDeleteRequest(BaseModel):
    name: str


@app.get("/health")
async def health() -> dict:
    return {"ok": True, "service": "qonduit-memory-gateway"}


@app.get("/v1/models")
@app.get("/models")
async def list_models() -> dict:
    async with httpx.AsyncClient(timeout=20.0) as client:
        r = await client.get(f"{LLAMA_BASE}/v1/models")
        if r.status_code >= 400:
            raise HTTPException(status_code=r.status_code, detail=r.text)
        data = r.json()
    return data


def get_request_user_id(request: Request) -> str:
    raw = request.headers.get("X-Qonduit-User", "").strip().lower()
    safe = "".join(c for c in raw if c.isalnum() or c in ("-", "_"))
    return safe or "default"


def ensure_upload_dir() -> None:
    os.makedirs(UPLOAD_DIR, exist_ok=True)


def sanitize_filename(filename: str) -> str:
    cleaned = filename.strip()
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", cleaned)
    cleaned = cleaned.strip("._")
    return cleaned or "uploaded_file"


def build_saved_upload_path(user_id: str, collection_name: str, filename: str) -> str:
    safe_user = sanitize_filename(user_id)
    safe_collection = sanitize_filename(collection_name)
    safe_filename = sanitize_filename(filename)
    file_id = str(uuid.uuid4())
    collection_dir = os.path.join(UPLOAD_DIR, safe_user, safe_collection)
    os.makedirs(collection_dir, exist_ok=True)
    return os.path.join(collection_dir, f"{file_id}__{safe_filename}")


def collection_upload_dir(user_id: str, collection_name: str) -> str:
    safe_user = sanitize_filename(user_id)
    safe_collection = sanitize_filename(collection_name)
    return os.path.join(UPLOAD_DIR, safe_user, safe_collection)


def delete_collection_upload_dir(user_id: str, collection_name: str) -> int:
    import shutil

    target_dir = collection_upload_dir(user_id, collection_name)
    if not os.path.exists(target_dir):
        return 0

    file_count = 0
    for _, _, files in os.walk(target_dir):
        file_count += len(files)

    shutil.rmtree(target_dir, ignore_errors=True)
    return file_count


def code_edit_artifact_dir(user_id: str) -> str:
    safe_user = sanitize_filename(user_id)
    target_dir = os.path.join(UPLOAD_DIR, safe_user, "__code_edits__")
    os.makedirs(target_dir, exist_ok=True)
    return target_dir


def save_code_edit_artifact(user_id: str, filename: str, content: str) -> dict[str, Any]:
    safe_name = sanitize_filename(Path(filename).name or "modified_file.txt")
    if not Path(safe_name).suffix:
        safe_name = f"{safe_name}.txt"

    artifact_id = str(uuid.uuid4())
    saved_path = os.path.join(code_edit_artifact_dir(user_id), f"{artifact_id}__{safe_name}")

    with open(saved_path, "w", encoding="utf-8") as f:
        f.write(content)

    return {
        "id": artifact_id,
        "name": safe_name,
        "saved_path": saved_path,
        "size_bytes": len(content.encode("utf-8")),
        "type": "code_edit_file",
    }


def chunk_text(text: str, chunk_size: int = 1200, overlap: int = 200) -> list[str]:
    cleaned = text.strip()
    if not cleaned:
        return []

    chunks: list[str] = []
    start = 0
    text_len = len(cleaned)

    while start < text_len:
        end = min(start + chunk_size, text_len)
        chunk = cleaned[start:end].strip()
        if chunk:
            chunks.append(chunk)
        if end >= text_len:
            break
        start = max(end - overlap, 0)

    return chunks


def extract_text_from_txt_like(file_path: str) -> str:
    with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
        return f.read()


def extract_text_from_pdf(file_path: str) -> str:
    reader = PdfReader(file_path)
    parts: list[str] = []

    for page in reader.pages:
        try:
            page_text = page.extract_text() or ""
        except Exception:
            page_text = ""
        if page_text.strip():
            parts.append(page_text)

    return "\n\n".join(parts)


def extract_text_from_docx(file_path: str) -> str:
    doc = Document(file_path)
    parts = [p.text for p in doc.paragraphs if p.text and p.text.strip()]
    return "\n\n".join(parts)


def extract_text_from_csv_file(file_path: str) -> str:
    rows: list[str] = []
    with open(file_path, "r", encoding="utf-8", errors="ignore", newline="") as f:
        reader = csv.reader(f)
        for row in reader:
            rows.append(" | ".join(cell.strip() for cell in row))
    return "\n".join(rows)


def extract_text_from_xlsx(file_path: str) -> str:
    wb = load_workbook(file_path, data_only=True)
    parts: list[str] = []

    for sheet in wb.worksheets:
        parts.append(f"# Sheet: {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            values = ["" if cell is None else str(cell) for cell in row]
            line = " | ".join(v.strip() for v in values if str(v).strip())
            if line:
                parts.append(line)

    return "\n".join(parts)


def extract_text_from_file(file_path: str, suffix: str) -> str:
    suffix = suffix.lower()

    if suffix in TEXT_EXTENSIONS:
        return extract_text_from_txt_like(file_path)

    if suffix == ".pdf":
        return extract_text_from_pdf(file_path)

    if suffix == ".docx":
        return extract_text_from_docx(file_path)

    if suffix == ".csv":
        return extract_text_from_csv_file(file_path)

    if suffix in {".xlsx", ".xls"}:
        return extract_text_from_xlsx(file_path)

    raise ValueError(f"Unsupported file type: {suffix}")


def latest_user_text(messages: list[ChatMessage]) -> str:
    for msg in reversed(messages):
        if msg.role == "user":
            return coerce_model_content_to_text(msg.content)
    return ""


def _debug_preview(text: str, limit: int = 280) -> str:
    preview = text.replace("\n", "\\n")
    if len(preview) > limit:
        return preview[:limit] + "...(truncated)"
    return preview


def debug_code_edit_event(event: str, **kwargs: Any) -> None:
    parts: list[str] = []
    for key, value in kwargs.items():
        try:
            rendered = str(value)
        except Exception:
            rendered = "<unrenderable>"
        parts.append(f"{key}={rendered}")
    print(f"[code-edit-debug] {event} | " + " | ".join(parts), flush=True)


def is_code_edit_request(text: str) -> bool:
    lowered = text.lower()
    return (
        "code edit" in lowered
        or "modified file" in lowered
        or "update this file" in lowered
        or "rewrite this file" in lowered
    )


def unwrap_structured_text_payload(text: str) -> str:
    raw = text.strip()
    if not raw.startswith("[") or "text" not in raw:
        return text

    parsed: Any | None = None
    try:
        parsed = json.loads(raw)
    except Exception:
        try:
            parsed = ast.literal_eval(raw)
        except Exception:
            return text

    if not isinstance(parsed, list):
        return text

    parts: list[str] = []
    for item in parsed:
        if isinstance(item, dict):
            kind = str(item.get("type", "")).lower()
            if kind in {"text", "input_text"}:
                value = item.get("text")
                if isinstance(value, str) and value.strip():
                    parts.append(value)

    if not parts:
        return text
    return "\n".join(parts)


def extract_requested_filename(text: str) -> str | None:
    patterns = [
        r"file:\s*([^\n\r]+)",
        r"for file:\s*([^\n\r]+)",
        r"filename:\s*([^\n\r]+)",
    ]
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.IGNORECASE)
        if match:
            value = match.group(1).strip()
            if value:
                return Path(value).name
    return None


def extract_inline_file_contents(text: str) -> str:
    normalized = unwrap_structured_text_payload(text)
    patterns = [
        r"Current file contents:\s*\n(?P<body>.*)$",
        r"Current file:\s*\n(?P<body>.*)$",
        r"<<<FILE\s*\n(?P<body>.*?)\nFILE\s*$",
    ]
    for pattern in patterns:
        match = re.search(pattern, normalized, flags=re.IGNORECASE | re.DOTALL)
        if match:
            body = match.group("body").strip("\n\r")
            if body.strip():
                return body
    return ""


def extract_code_edit_instruction(text: str) -> str:
    working = unwrap_structured_text_payload(text)

    working = re.sub(
        r"^\s*Code edit request for file:\s*[^\n\r]+\s*",
        "",
        working,
        flags=re.IGNORECASE,
    )
    working = re.sub(
        r"Current file contents:\s*\n.*$",
        "",
        working,
        flags=re.IGNORECASE | re.DOTALL,
    )
    working = re.sub(
        r"Current file:\s*\n.*$",
        "",
        working,
        flags=re.IGNORECASE | re.DOTALL,
    )
    working = re.sub(
        r"<<<FILE\s*\n.*?\nFILE\s*$",
        "",
        working,
        flags=re.IGNORECASE | re.DOTALL,
    )

    return working.strip()


def build_code_edit_model_input(
    requested_filename: str | None,
    instruction: str,
    file_contents: str,
) -> str:
    target_name = requested_filename or "modified_file.txt"
    safe_instruction = instruction.strip() or "Update the file as requested."

    return (
        f"Target file: {target_name}\n\n"
        f"Edit instruction:\n{safe_instruction}\n\n"
        "Current file contents:\n"
        f"{file_contents}"
    )


def build_code_edit_contract_system_prompt(requested_filename: str | None) -> str:
    target_name = requested_filename or "modified_file.txt"
    return (
        "You are handling a code edit request. "
        "You will be given the target file name, a concrete edit instruction, and the current file contents. "
        "Use the provided file contents as the source of truth. "
        "Do not return a unified diff. "
        "Do not return markdown fences. "
        "Do not include conversational filler. "
        "Return only valid JSON with this exact schema:\n"
        "{\n"
        '  "executive_summary": ["short bullet", "short bullet"],\n'
        '  "change_summary": ["technical bullet", "technical bullet"],\n'
        '  "patch_confidence": "high|medium|low",\n'
        '  "modified_file": {\n'
        f'    "name": "{target_name}",\n'
        '    "content": "full updated file contents here"\n'
        "  }\n"
        "}\n"
        "Rules:\n"
        "1. executive_summary must be short and human-readable.\n"
        "2. change_summary must be technical and concise.\n"
        "3. patch_confidence must be exactly one of: high, medium, low.\n"
        "4. modified_file.content must contain the full updated file contents.\n"
        "5. Never return a diff.\n"
        "6. Never omit modified_file.content if you can complete the edit.\n"
        "7. Only use low patch_confidence when the instruction is ambiguous or the supplied file contents are clearly insufficient.\n"
    )


def coerce_model_content_to_text(content: Any) -> str:
    if isinstance(content, str):
        return content

    if isinstance(content, list):
        parts: list[str] = []
        for item in content:
            if isinstance(item, dict):
                text = item.get("text")
                if isinstance(text, str) and text.strip():
                    parts.append(text)
            elif isinstance(item, str) and item.strip():
                parts.append(item)
        return "\n".join(parts).strip()

    if isinstance(content, dict):
        text = content.get("text")
        if isinstance(text, str):
            return text

    return str(content or "")


def extract_json_object(raw: str) -> str | None:
    start = raw.find("{")
    end = raw.rfind("}")
    if start == -1 or end == -1 or end <= start:
        return None
    return raw[start:end + 1]


def normalize_summary_lines(value: Any) -> list[str]:
    if isinstance(value, list):
        items = []
        for item in value:
            text = str(item).strip()
            if text:
                items.append(text)
        return items

    if isinstance(value, str):
        lines = []
        for line in value.splitlines():
            cleaned = line.strip().lstrip("-•* ").strip()
            if cleaned:
                lines.append(cleaned)
        return lines

    return []


def normalize_patch_confidence(value: Any) -> str:
    text = str(value or "").strip().lower()
    if text in {"high", "medium", "low"}:
        return text
    return "low"


def extract_fenced_code_block(raw: str) -> str:
    blocks = re.findall(
        r"```(?:[a-zA-Z0-9_+\-#.]+)?\s*\n(.*?)```",
        raw,
        flags=re.DOTALL,
    )
    cleaned_blocks = [block.strip("\n\r") for block in blocks if block.strip()]
    if not cleaned_blocks:
        return ""

    cleaned_blocks.sort(key=len, reverse=True)
    return cleaned_blocks[0]


def build_non_json_code_edit_fallback(
    raw: str,
    requested_filename: str | None,
) -> dict[str, Any] | None:
    recovered = extract_fenced_code_block(raw)
    if not recovered.strip():
        return None

    fallback_name = requested_filename or "modified_file.txt"
    return {
        "executive_summary": [
            "Recovered a code-edit result from a non-JSON model response."
        ],
        "change_summary": [
            "The model did not follow the JSON contract, so the gateway extracted "
            "the largest fenced code block as the modified file content."
        ],
        "patch_confidence": "low",
        "modified_file": {
            "name": Path(fallback_name).name or "modified_file.txt",
            "content": recovered,
        },
    }


def parse_code_edit_response(raw: Any, requested_filename: str | None) -> dict[str, Any]:
    raw_text = coerce_model_content_to_text(raw)
    default_name = requested_filename or "modified_file.txt"
    default = {
        "executive_summary": [
            "The model did not return the requested structured code-edit response."
        ],
        "change_summary": [
            "No attachment-ready modified file was produced from the response."
        ],
        "patch_confidence": "low",
        "modified_file": {
            "name": default_name,
            "content": "",
        },
    }

    blob = extract_json_object(raw_text)
    if not blob:
        recovered = build_non_json_code_edit_fallback(raw_text, requested_filename)
        if recovered is not None:
            return recovered
        return default

    try:
        parsed = json.loads(blob)
    except Exception:
        recovered = build_non_json_code_edit_fallback(raw_text, requested_filename)
        if recovered is not None:
            return recovered
        return default

    executive_summary = normalize_summary_lines(parsed.get("executive_summary"))
    change_summary = normalize_summary_lines(parsed.get("change_summary"))
    patch_confidence = normalize_patch_confidence(parsed.get("patch_confidence"))

    modified_file = parsed.get("modified_file", {})
    if not isinstance(modified_file, dict):
        modified_file = {}

    modified_name = str(
        modified_file.get("name")
        or parsed.get("modified_file_name")
        or default_name
    ).strip() or default_name

    modified_content = str(
        modified_file.get("content")
        or parsed.get("modified_file_content")
        or ""
    )

    if not modified_content.strip():
        recovered_content = extract_fenced_code_block(raw_text)
        if recovered_content.strip():
            modified_content = recovered_content
            if not change_summary:
                change_summary = [
                    "Recovered modified file content from a fenced code block "
                    "because modified_file.content was empty."
                ]
            patch_confidence = "low"

    if not executive_summary:
        executive_summary = ["Prepared a code-edit response."]
    if not change_summary:
        if modified_content.strip():
            change_summary = [f"Prepared updated file contents for {modified_name}."]
        else:
            change_summary = ["No updated file contents were provided by the model."]

    return {
        "executive_summary": executive_summary,
        "change_summary": change_summary,
        "patch_confidence": patch_confidence,
        "modified_file": {
            "name": Path(modified_name).name or default_name,
            "content": modified_content,
        },
    }


def format_code_edit_summary(parsed: dict[str, Any], artifact: dict[str, Any] | None) -> str:
    lines: list[str] = []

    lines.append("Executive Summary")
    for item in parsed["executive_summary"]:
        lines.append(f"• {item}")

    lines.append("")
    lines.append("Change Summary")
    for item in parsed["change_summary"]:
        lines.append(f"• {item}")

    lines.append("")
    lines.append(f"Patch Confidence: {str(parsed['patch_confidence']).upper()}")

    if artifact is not None:
        lines.append("")
        lines.append(f"Prepared File: {artifact['name']}")
        lines.append(f"Saved Path: {artifact['saved_path']}")

    return "\n".join(lines).strip()


def sse_chunk(model: str, content: str = "", finish_reason: str | None = None) -> str:
    payload = {
        "id": f"chatcmpl-qonduit-{uuid.uuid4().hex}",
        "object": "chat.completion.chunk",
        "created": int(time.time()),
        "model": model,
        "choices": [
            {
                "index": 0,
                "delta": {"content": content} if content else {},
                "finish_reason": finish_reason,
            }
        ],
    }
    return f"data: {json.dumps(payload, ensure_ascii=False)}\n\n"


def split_for_stream(text: str, chunk_size: int = 180) -> list[str]:
    if not text:
        return []
    parts: list[str] = []
    start = 0
    while start < len(text):
        parts.append(text[start:start + chunk_size])
        start += chunk_size
    return parts


def strip_markdown_fences(text: str) -> str:
    trimmed = text.strip()
    if not trimmed.startswith("```"):
        return trimmed

    match = re.match(r"^```[^\n]*\n(?P<body>.*)\n```$", trimmed, flags=re.DOTALL)
    if not match:
        return trimmed
    return match.group("body").strip("\n\r")


def try_apply_extension_edit_locally(instruction: str, original_file: str) -> str | None:
    lowered = instruction.lower()
    if "add" not in lowered or "extension" not in lowered:
        return None

    ext_match = re.search(r"(\.[a-z0-9_+-]+)", instruction, flags=re.IGNORECASE)
    if not ext_match:
        return None
    extension = ext_match.group(1).lower()

    collection_match = re.search(
        r"(?P<prefix>\b[A-Z_]*EXTENSIONS\b\s*=\s*)(?P<open>[\[\(\{])(?P<body>.*?)(?P<close>[\]\)\}])",
        original_file,
        flags=re.DOTALL,
    )
    if not collection_match:
        return None

    body = collection_match.group("body")
    if re.search(rf"['\"]{re.escape(extension)}['\"]", body, flags=re.IGNORECASE):
        return original_file

    quote = "'" if "'" in body else '"'
    updated_body = body.rstrip()
    if updated_body and not updated_body.endswith(","):
        updated_body = f"{updated_body},"

    insertion = f"\n    {quote}{extension}{quote},"
    replaced = (
        f"{collection_match.group('prefix')}"
        f"{collection_match.group('open')}"
        f"{updated_body}{insertion}\n"
        f"{collection_match.group('close')}"
    )

    start, end = collection_match.span()
    return f"{original_file[:start]}{replaced}{original_file[end:]}"


def looks_like_full_file_content(original_file: str, candidate_file: str) -> bool:
    candidate = candidate_file.strip()
    if not candidate:
        return False

    if len(candidate) < max(120, int(len(original_file) * 0.2)):
        return False

    if "\n" not in candidate and len(original_file) > 400:
        return False

    return True


async def recover_modified_file_with_retry(
    model: str,
    requested_filename: str | None,
    instruction: str,
    original_file_contents: str,
) -> str:
    target_name = requested_filename or "modified_file.txt"
    retry_messages = [
        {
            "role": "system",
            "content": (
                "You are repairing a failed code-edit response. "
                "Return only the full updated file contents. "
                "Do not return JSON. "
                "Do not return markdown fences. "
                "Do not explain anything."
            ),
        },
        {
            "role": "user",
            "content": (
                f"Target file: {target_name}\n\n"
                f"Edit instruction:\n{instruction.strip()}\n\n"
                "Current file contents:\n"
                f"{original_file_contents}"
            ),
        },
    ]

    retry_payload = {
        "model": model,
        "messages": retry_messages,
        "max_tokens": 8192,
        "temperature": 0.0,
        "stream": False,
    }

    async with httpx.AsyncClient(timeout=45.0) as client:
        retry_response = await client.post(
            f"{LLAMA_BASE}/v1/chat/completions",
            json=retry_payload,
        )
        retry_response.raise_for_status()
        retry_data = retry_response.json()

    raw_content = str(retry_data["choices"][0]["message"]["content"] or "")
    return strip_markdown_fences(raw_content)


@app.post("/rag/test-ingest")
async def rag_test_ingest(req: RagIngestRequest, request: Request) -> dict:
    user_id = get_request_user_id(request)
    doc_id = await add_document(
        text=req.text,
        metadata={
            "source": req.source,
            "collection": req.collection,
            "document_name": req.document_name,
        },
        user_id=user_id,
    )
    return {"ok": True, "id": doc_id}


@app.post("/rag/test-search")
async def rag_test_search(req: RagSearchRequest, request: Request) -> dict:
    user_id = get_request_user_id(request)
    results = await search_documents(
        req.query,
        limit=req.limit,
        collection=req.collection,
        user_id=user_id,
    )
    return {"ok": True, "results": results}


@app.get("/rag/collections")
async def rag_list_collections(request: Request) -> dict:
    user_id = get_request_user_id(request)
    return {"ok": True, "collections": list_collections(user_id=user_id)}


@app.post("/rag/collections/create")
async def rag_create_collection(req: RagCollectionCreateRequest, request: Request) -> dict:
    user_id = get_request_user_id(request)
    name = req.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Collection name cannot be empty")

    await create_collection_marker(name, user_id=user_id)
    return {"ok": True, "collection": name}


@app.post("/rag/collections/delete")
async def rag_delete_collection(req: RagCollectionDeleteRequest, request: Request) -> dict:
    user_id = get_request_user_id(request)
    collection_name = req.name.strip()
    if not collection_name:
        raise HTTPException(status_code=400, detail="Collection name cannot be empty")

    deleted_points = 0

    try:
        points, _ = qdrant.scroll(
            collection_name=COLLECTION_NAME,
            scroll_filter=Filter(
                must=[
                    FieldCondition(
                        key="collection",
                        match=MatchValue(value=collection_name),
                    ),
                    FieldCondition(
                        key="user_id",
                        match=MatchValue(value=user_id),
                    ),
                ]
            ),
            limit=10000,
            with_payload=False,
            with_vectors=False,
        )

        point_ids = [p.id for p in points if p.id is not None]
        if point_ids:
            qdrant.delete(
                collection_name=COLLECTION_NAME,
                points_selector=point_ids,
            )
            deleted_points = len(point_ids)

        deleted_files = delete_collection_upload_dir(user_id, collection_name)

        return {
            "ok": True,
            "collection": collection_name,
            "deleted_points": deleted_points,
            "deleted_files": deleted_files,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to delete collection: {e}")


@app.post("/rag/upload")
async def rag_upload_document(
    request: Request,
    file: UploadFile = File(...),
    collection: str = Form(...),
    source: str = Form("qonduit_upload"),
) -> dict:
    ensure_upload_dir()

    user_id = get_request_user_id(request)
    collection_name = collection.strip()
    if not collection_name:
        raise HTTPException(status_code=400, detail="Collection cannot be empty")

    filename = file.filename or "uploaded_file"
    filename = sanitize_filename(filename)
    suffix = Path(filename).suffix.lower()

    if not suffix:
        raise HTTPException(status_code=400, detail="Uploaded file must have an extension")

    saved_path = build_saved_upload_path(user_id, collection_name, filename)

    try:
        with open(saved_path, "wb") as f:
            content = await file.read()
            f.write(content)

        extracted_text = extract_text_from_file(saved_path, suffix)
        if not extracted_text.strip():
            raise HTTPException(status_code=400, detail="No readable text found in file")

        chunks = chunk_text(extracted_text)
        if not chunks:
            raise HTTPException(status_code=400, detail="No text chunks generated from file")

        chunk_ids: list[str] = []
        for idx, chunk in enumerate(chunks):
            chunk_id = await add_document(
                text=chunk,
                metadata={
                    "source": source,
                    "collection": collection_name,
                    "document_name": filename,
                    "chunk_index": idx,
                    "file_type": suffix,
                    "saved_path": saved_path,
                },
                user_id=user_id,
            )
            chunk_ids.append(chunk_id)

        return {
            "ok": True,
            "collection": collection_name,
            "document_name": filename,
            "chunks_added": len(chunk_ids),
            "saved_path": saved_path,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to upload document: {e}")


@app.get("/v1/chat/completions")
@app.get("/chat/completions")
async def chat_completions_help() -> dict:
    return {
        "ok": True,
        "message": "Use POST /v1/chat/completions (or /chat/completions) with a JSON body.",
    }


@app.post("/v1/chat/completions")
@app.post("/chat/completions")
async def chat(req: GatewayChatRequest, request: Request) -> Any:
    user_id = get_request_user_id(request)
    context_size = req.resolved_context_size()
    conversation_id = (
        (req.conversation_id or "").strip()
        or (request.headers.get("X-Qonduit-Conversation", "").strip())
        or (req.user or "").strip()
        or "default"
    )
    state = load_conversation(conversation_id)

    prior_recent = state.get("recent_messages", [])
    summary = state.get("summary", "")

    incoming = [
        {
            "role": m.role,
            "content": coerce_model_content_to_text(m.content),
        }
        for m in req.messages
    ]
    combined_recent = prior_recent + incoming

    trimmed_recent, _ = trim_recent_messages(
        combined_recent,
        summary=summary,
        system_prompt=DEFAULT_SYSTEM_PROMPT,
        context_size=context_size,
    )

    overflow_count = len(combined_recent) - len(trimmed_recent)

    if overflow_count > 0:
        older = combined_recent[:overflow_count]
        summary = await summarize_messages(req.model, summary, older)

    remaining_recent = combined_recent[overflow_count:]
    trimmed_recent, prompt_tokens = trim_recent_messages(
        remaining_recent,
        summary=summary,
        system_prompt=DEFAULT_SYSTEM_PROMPT,
        context_size=context_size,
    )

    budget = build_budget(context_size)
    max_tokens = min(req.max_tokens, budget.reserved_output)

    latest_text = latest_user_text(req.messages)

    rag_results = []
    rag_chunks = []

    if latest_text.strip():
        try:
            rag_results = await search_documents(
                latest_text,
                limit=2,
                collection=req.rag_collection or conversation_id,
                user_id=user_id,
            )
            rag_chunks = [
                item["text"].strip()
                for item in rag_results
                if item.get("text", "").strip()
            ]
        except Exception:
            rag_results = []
            rag_chunks = []

    rag_context = "\n\n".join(rag_chunks)

    final_messages = [
        {"role": "system", "content": DEFAULT_SYSTEM_PROMPT},
        {"role": "system", "content": f"Rolling summary:\n{summary or '(none)'}"},
    ]

    if rag_context:
        final_messages.append(
            {
                "role": "system",
                "content": (
                    "Relevant retrieved knowledge:\n"
                    f"{rag_context}"
                ),
            }
        )

    final_messages.extend(trimmed_recent)

    state["summary"] = summary
    state["recent_messages"] = trimmed_recent[-8:]
    state["last_model"] = req.model
    state["last_context_size"] = context_size
    state["last_prompt_tokens"] = prompt_tokens
    state["last_reserved_output"] = budget.reserved_output
    save_conversation(conversation_id, state)

    payload = {
        "model": req.model,
        "messages": final_messages,
        "max_tokens": max_tokens,
        "temperature": req.temperature,
        "stream": req.stream,
    }

    async def event_stream():
        stream_payload = dict(payload)
        stream_payload["stream"] = True
        stream_start = time.perf_counter()
        emitted_chunks = 0
        emitted_chars = 0
        assistant_parts: list[str] = []
        sent_done = False
        timeout = httpx.Timeout(
            connect=UPSTREAM_CONNECT_TIMEOUT_SECONDS,
            read=None,
            write=UPSTREAM_WRITE_TIMEOUT_SECONDS,
            pool=UPSTREAM_POOL_TIMEOUT_SECONDS,
        )

        try:
            async with httpx.AsyncClient(timeout=timeout) as client:
                async with client.stream(
                    "POST",
                    f"{LLAMA_BASE}/v1/chat/completions",
                    json=stream_payload,
                ) as r:
                    upstream_ms = int((time.perf_counter() - stream_start) * 1000)
                    logger.info(
                        "stream_upstream_response conversation_id=%s model=%s "
                        "status=%s latency_ms=%s",
                        conversation_id,
                        req.model,
                        r.status_code,
                        upstream_ms,
                    )

                    if r.status_code >= 400:
                        upstream_error = await r.aread()
                        error_text = upstream_error.decode(
                            "utf-8",
                            errors="replace",
                        )
                        yield sse_chunk(
                            req.model,
                            content=f"Upstream error ({r.status_code}): "
                            f"{error_text}",
                        )
                        yield sse_chunk(req.model, finish_reason="stop")
                        yield "data: [DONE]\n\n"
                        return

                    upstream_lines = r.aiter_lines()
                    while True:
                        try:
                            line = await asyncio.wait_for(
                                anext(upstream_lines),
                                timeout=STREAM_KEEPALIVE_INTERVAL_SECONDS,
                            )
                        except asyncio.TimeoutError:
                            yield ": keep-alive\n\n"
                            continue
                        except StopAsyncIteration:
                            break

                        if not line:
                            continue
                        if not line.startswith("data:"):
                            continue

                        data_line = line[5:].strip()
                        if not data_line:
                            continue

                        if data_line == "[DONE]":
                            sent_done = True
                            yield "data: [DONE]\n\n"
                            break

                        emitted_chunks += 1
                        yield f"data: {data_line}\n\n"

                        try:
                            parsed = json.loads(data_line)
                        except json.JSONDecodeError:
                            continue

                        choices = parsed.get("choices", [])
                        if not choices:
                            continue

                        delta = choices[0].get("delta", {})
                        text_delta = coerce_model_content_to_text(
                            delta.get("content", ""),
                        )
                        if text_delta:
                            emitted_chars += len(text_delta)
                            assistant_parts.append(text_delta)

                    if not sent_done:
                        yield sse_chunk(req.model, finish_reason="stop")
                        yield "data: [DONE]\n\n"
        except Exception as e:
            logger.exception(
                "stream_request_failed conversation_id=%s model=%s error=%s",
                conversation_id,
                req.model,
                str(e),
            )
            yield sse_chunk(req.model, content=f"Gateway streaming error: {str(e)}")
            yield sse_chunk(req.model, finish_reason="stop")
            yield "data: [DONE]\n\n"
            return

        assistant_content = "".join(assistant_parts)
        logger.info(
            "stream_complete conversation_id=%s model=%s chunks=%s chars=%s",
            conversation_id,
            req.model,
            emitted_chunks,
            emitted_chars,
        )

        assistant_message = {
            "role": "assistant",
            "content": assistant_content,
        }

        state["summary"] = summary
        state["recent_messages"] = (trimmed_recent + [assistant_message])[-8:]
        state["last_model"] = req.model
        state["last_context_size"] = context_size
        state["last_prompt_tokens"] = prompt_tokens
        state["last_reserved_output"] = budget.reserved_output
        save_conversation(conversation_id, state)

    if req.stream:
        logger.info(
            "chat_request stream=true conversation_id=%s model=%s messages=%s",
            conversation_id,
            req.model,
            len(req.messages),
        )
        return StreamingResponse(
            event_stream(),
            media_type="text/event-stream",
            headers={
                "Cache-Control": "no-cache",
                "Connection": "keep-alive",
            },
        )

    non_stream_start = time.perf_counter()
    async with httpx.AsyncClient(timeout=120.0) as client:
        r = await client.post(f"{LLAMA_BASE}/v1/chat/completions", json=payload)
        if r.status_code >= 400:
            raise HTTPException(status_code=r.status_code, detail=r.text)
        data = r.json()
    non_stream_ms = int((time.perf_counter() - non_stream_start) * 1000)
    logger.info(
        "chat_request stream=false conversation_id=%s model=%s latency_ms=%s messages=%s",
        conversation_id,
        req.model,
        non_stream_ms,
        len(req.messages),
    )

    message_content = (
        data.get("choices", [{}])[0]
        .get("message", {})
        .get("content", "")
    )
    assistant_content = coerce_model_content_to_text(message_content)

    assistant_message = {
        "role": "assistant",
        "content": assistant_content,
    }

    state["summary"] = summary
    state["recent_messages"] = (trimmed_recent + [assistant_message])[-8:]
    state["last_model"] = req.model
    state["last_context_size"] = context_size
    state["last_prompt_tokens"] = prompt_tokens
    state["last_reserved_output"] = budget.reserved_output
    save_conversation(conversation_id, state)

    return {
        "id": data.get("id", f"chatcmpl-qonduit-{uuid.uuid4().hex}"),
        "object": "chat.completion",
        "created": data.get("created", int(time.time())),
        "model": data.get("model", req.model),
        "choices": [
            {
                "index": 0,
                "message": {
                    "role": "assistant",
                    "content": assistant_content,
                },
                "finish_reason": (
                    data.get("choices", [{}])[0].get("finish_reason") or "stop"
                ),
            }
        ],
        "usage": data.get("usage", {}),
    }
